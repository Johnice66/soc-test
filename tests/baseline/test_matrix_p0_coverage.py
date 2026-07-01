"""确保 Excel P0 用例均已实现，或已登记明确的暂缓原因。"""
import ast
from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

CASE_ID = "SOC-MATRIX-001"
MITRE = "N/A"
pytestmark = [pytest.mark.p0, pytest.mark.baseline]

ROOT = Path(__file__).resolve().parents[2]


def _implemented_case_ids() -> set[str]:
    case_ids = set()
    for path in (ROOT / "tests").rglob("test_*.py"):
        tree = ast.parse(path.read_text(encoding="utf-8"), filename=str(path))
        for node in tree.body:
            if not isinstance(node, (ast.Assign, ast.AnnAssign)):
                continue
            targets = node.targets if isinstance(node, ast.Assign) else [node.target]
            if not any(isinstance(target, ast.Name) and target.id == "CASE_ID" for target in targets):
                continue
            value = node.value
            if isinstance(value, ast.Constant) and isinstance(value.value, str):
                case_ids.add(value.value)
    return case_ids


def test_all_p0_cases_are_implemented_or_deferred(evidence_recorder):
    rec = evidence_recorder
    deferred_doc = yaml.safe_load((ROOT / "config/deferred_cases.yaml").read_text(encoding="utf-8"))
    workbook = load_workbook(
        ROOT / "AI-SOC测试点与测试用例矩阵_MITRE_ATTCK映射版.xlsx",
        read_only=True,
        data_only=True,
    )
    rows = workbook["用例矩阵"].iter_rows(values_only=True)
    header = {name: index for index, name in enumerate(next(rows))}
    p0 = {
        row[header["用例ID"]]
        for row in rows
        if row[header["用例ID"]] and row[header["优先级"]] == "P0"
    }
    workbook.close()
    implemented = _implemented_case_ids() & p0
    deferred = set((deferred_doc or {}).get("cases", {}))

    unclassified = p0 - implemented - deferred
    overlap = implemented & deferred
    unknown_deferred = deferred - p0
    rec.application({
        "p0_total": len(p0),
        "implemented": len(implemented),
        "deferred": len(deferred),
        "unclassified": sorted(unclassified),
        "overlap": sorted(overlap),
        "unknown_deferred": sorted(unknown_deferred),
    })
    rec.assertion("all_p0_classified", not unclassified, str(sorted(unclassified)))
    rec.assertion("implemented_not_deferred", not overlap, str(sorted(overlap)))
    rec.assertion("deferred_ids_exist", not unknown_deferred, str(sorted(unknown_deferred)))
    assert not unclassified
    assert not overlap
    assert not unknown_deferred
    rec.finish("PASS", f"P0={len(p0)}：已实现 {len(implemented)}，条件不足暂缓 {len(deferred)}")
