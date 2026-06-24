#!/usr/bin/env python3
"""把 AI-SOC测试点与测试用例矩阵_MITRE_ATTCK映射版.xlsx 转成 config/test_matrix.yaml

用法：
  python scripts/xlsx_to_yaml.py
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "AI-SOC测试点与测试用例矩阵_MITRE_ATTCK映射版.xlsx"
OUT = ROOT / "config" / "test_matrix.yaml"


def main() -> int:
    if not XLSX.exists():
        print(f"未找到 xlsx：{XLSX}", file=sys.stderr)
        return 1
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["用例矩阵"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {h: i for i, h in enumerate(header)}

    keys_map = {
        "category": "一级分类",
        "scenario": "测试场景",
        "tactic": "ATT&CK战术",
        "tech_id": "ATT&CK技术ID",
        "tech_name": "ATT&CK技术名称",
        "mitre_mainline": "是否MITRE主线",
        "soc_domain": "AI-SOC功能域",
        "data_source": "数据源/日志源",
        "precondition": "前置条件",
        "steps": "测试步骤（摘要）",
        "expected_platform": "预期平台/业务结果",
        "expected_wazuh": "预期Wazuh检测",
        "expected_vigil_ai": "预期Vigil/AI研判",
        "expected_response": "预期响应/闭环",
        "strength": "强度等级",
        "priority": "优先级",
        "acceptance": "验收指标",
        "source_url": "来源URL",
    }

    cases: dict = {}
    for r in rows[1:]:
        cid = r[idx["用例ID"]]
        if not cid:
            continue
        meta = {}
        for k, col in keys_map.items():
            v = r[idx[col]] if col in idx else None
            meta[k] = str(v) if v is not None else ""
        cases[str(cid)] = meta

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        yaml.safe_dump(
            {"source": str(XLSX.name), "total": len(cases), "cases": cases},
            f, allow_unicode=True, sort_keys=False, width=120,
        )
    print(f"写入 {OUT}（{len(cases)} 条用例）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
